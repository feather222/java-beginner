#datetime库，openpyxl库导入
import datetime
import openpyxl
#加载工作簿
wb = openpyxl.load_workbook('阿里巴巴2020年股票数据.xlsx')
#获取工作表名称
print(wb.sheetnames)
#获取第一个工作表 --> Worksheet
sheet = wb.worksheets[0]
#获取单元格范围
print(sheet.dimensions)
#获取行数和列数
print(sheet.max_row,sheet.max_column)
#获取具体单元格的value
print(sheet.cell(3,3).value)
print(sheet['C3'].value)
print(sheet['G252'].value)
#获取多个单元格的值
print(sheet['A2:G5'])
#获取所有单元格的值
for row_ch in range(2,sheet.max_row+1):
    for col_ch in 'ABCDEFG':
        value = sheet[f'{col_ch}{row_ch}'].value    #注意行和列位置问题，先列后行
        if type(value) == datetime.datetime:
            print(value.strftime('%Y年%m月%d日'),end='\t')
        elif type(value) == int:
            print(f'{value:<10d}',end='\t')
        elif type(value) == float:
            print(f'{value:.4f}',end='\t')
        else:
            print(value,end='\t')
    print()

#写入excel文件
import random
import openpyxl
#创建工作簿
wb = openpyxl.Workbook()
#添加工作表
sheet = wb.active
sheet.title = '期末成绩'
#添加表头titles
titles = ['name','chinese','math','english']
for index, title in enumerate(titles):
    sheet.cell(1, index+1,title)
#在第一列添加names（注意：从第二行开始）
names = ['one','two','three','four','five']
for row_index, name in enumerate(names):
    sheet.cell(row_index + 2, 1,name)
    #写入成绩
    for col_index in range(2,5):
        sheet.cell(row_index + 2, col_index, random.randrange(50,101))
#保存文件
wb.save('my_scores.xlsx')

#调整样式和公式计算
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
#对齐方式
alignment = Alignment(horizontal='center', vertical = 'center')
#边框线条
side = Side(color='ff7f50', style = 'mediumDashed')
#加载工作簿
wb = openpyxl.load_workbook('my_scores.xlsx')
#获取第一个工作表
sheet = wb.worksheets[0]
#调整第一行的行高和第E列的列宽
sheet.row_dimensions[1].height = 30
sheet.column_dimensions['E'].width = 12
sheet['E1'] = '平均分'
#设置字体 bold是加粗
sheet.cell(1, 5).font = Font(size= 18, bold= True, color= 'ff1493', name= '华文楷体')
#设置对齐方式
sheet.cell(1,5).alignment = alignment
#设置单元格边框
sheet.cell(1,5).border = Border(left= side, right= side, top= side, bottom= side)
for i in range(2,7):
    #公式计算每个学生的平均分
    sheet[f'E{i}'] = f'= average(B{i}:D{i})'
    sheet.cell(i,5).font = Font(size= 12, color= '4169e1', italic= True)
    sheet.cell(i,5).alignment = alignment
wb.save('我的成绩.xlsx')

#生成统计图表
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
#创建只能写入权限工作簿
wb = Workbook(write_only = True)
#创建工作表
sheet = wb.create_sheet()
rows = [
    ('类别', '销售A组', '销售B组'),
    ('手机', 40, 30),
    ('平板', 50, 60),
    ('笔记本', 80, 70),
    ('外围设备', 20, 10),
]
#向表单中添加行
for row in rows:
    sheet.append(row)
#创建图表对象
chart = BarChart()
chart.type = 'col'      #垂直柱状图
chart.style = 10        #预设样式
#设置图标的标题
chart.title = '销售统计图'
#设置图表纵轴的标题
chart.y_axis.title = '销量'
#设置图表横轴的标题
chart.x_axis.title = '商品类别'
#设置数据的范围
data = Reference(sheet, min_col = 2, max_col = 3, min_row = 1, max_row = 5)
#设置分类的范围
cats = Reference(sheet, min_row = 2, max_row = 5, min_col= 1)
#将数据添加到表格
chart.add_data(data, titles_from_data= True)
#给图表设置分类
chart.set_categories(cats)
chart.shape = 4
#将图表加到表单指定的单元格中
sheet.add_chart(chart, 'A10')
wb.save('demo.xlsx')
